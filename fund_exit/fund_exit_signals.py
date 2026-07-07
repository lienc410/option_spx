# -*- coding: utf-8 -*-
"""基金技术面分批清仓信号工具 (Strategy Spec v3 — 2026-06 审计后重构)

定位：纪律 / 行为约束工具，非 alpha 择时、非投资建议。
详见 task/fund_exit_strategy_spec.md。

核心模型：clip = base(周频截止日 TWAP·强制) + extra(弱倾斜前置·可被超卖否决)
  · base  ：非强势仓每周卖 1/剩余周数(of 现值) → 线性清空至 deadline。操作周频。
  · extra ：弱势(③④⑤)前置 = base × min(LEAN_SLOPE×回撤, LEAN_MAX)；RSI<30 时归零，base 不变。
  · 强势①(上升趋势)纯持有，base=extra=0；破势后才入清仓钟。

规则（优先级取首个命中；深套/追踪统一用"滚动高"参照）：
  ① 上升趋势 : 最新>MA20>MA60                       -> 纯持有(base=extra=0)
  ② 深套     : 下降 且 距滚动高 <= -DEEP             -> 仅 base TWAP（不前置砸底）
  ③ 确认下降 : 下降 且 距滚动高 > -DEEP              -> base TWAP + 弱倾斜前置
  ④ 震荡破带 : 现价 <= 滚动高*(1-TRAIL_σ)            -> base TWAP + 弱倾斜前置
  ⑤ 破MA20   : 震荡 且 MA20>MA60 且 现价<MA20        -> base TWAP + 轻前置
  ⑥ 其余     : -> 仅 base TWAP

超卖否决 = 仅 RSI<RSI_OVERSOLD（审计 A1：去掉"创新低"——那是下行动能、非超卖）。
TRAIL σ-scaled(~1σ 6周, clip 5-14%)；趋势 band-hysteresis(±1%) 压日度 whipsaw。
"""
from __future__ import annotations

import math
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager

# ── CJK 字体（macOS）：优先全覆盖的 Arial Unicode，避免简体缺字 ──────
for _path, _name in [("/Library/Fonts/Arial Unicode.ttf", "Arial Unicode MS"),
                     ("/System/Library/Fonts/Supplemental/Arial Unicode.ttf", "Arial Unicode MS"),
                     ("/System/Library/Fonts/STHeiti Medium.ttc", "Heiti TC")]:
    try:
        font_manager.fontManager.addfont(_path)
        plt.rcParams["font.sans-serif"] = [_name]
        break
    except Exception:
        continue
plt.rcParams["axes.unicode_minus"] = False

import akshare as ak

OUTDIR = Path(__file__).resolve().parent

# ── 参数（2026-06 周频截止日 TWAP 重构）────────────────────────────
# 清仓节奏 clip = base(周频截止日 TWAP·强制) + extra(弱倾斜前置·可被超卖否决)
#   · base ：每只非强势仓每周卖 1/剩余周数(of 现值) → 线性清空至 deadline（无状态也按期清完）。
#            操作周频；超卖只挡 extra、不挡 base（base 仍走，避免砸在底但保持按期）。
#   · extra：弱势(③④⑤)按回撤前置 = base × min(LEAN_SLOPE×回撤, LEAN_MAX)；②深套/⑥/⓪不前置。
#   · 强势①(上升趋势)纯持有，base=extra=0（PM 接受 ~60% 惰性，破势后才入清仓钟）。
DEEP = 0.15               # 深套界：距滚动高 ≤ -DEEP（A6：深套/追踪统一用滚动高）
RSI_OVERSOLD = 30         # 超卖否决阈
DEADLINE_FAST = date(2026, 7, 16)   # 可减持(非上升)仓加速清空的节奏锚点
DEADLINE_SLOW = date(2026, 8, 31)   # 上升趋势仓 延长清仓（慢速 TWAP，给赢家更多时间）
# ── 2026-07-06 PM: deadline 软化 + 回前高减仓 ─────────────────────
# 软化: fast 池周数下限 SOFT_FLOOR_WEEKS → 过期不悬崖到 100%/周, 维持 1/4=25%/周持续出清,
#       给"回前高"反弹留窗口(前高=数个交易日前的滚动高, 锚新鲜)。
# 回前高(⑦, 优先于①): 距滚动高回到 -REC_BAND 内 且 近 REC_LOOKBACK 日内曾破势(非上升)
#       → 减 REC_CLIP/周, 反弹到哪卖到哪; 即使反弹强到重回上升也不转慢池(否则反弹越强越不卖)。
#       持续上升从未破势者(真赢家)不触发, 仍走延长慢清。
SOFT_FLOOR_WEEKS = 4
REC_BAND = 0.03
REC_CLIP = 0.50
REC_LOOKBACK = 15
LEAN_SLOPE = 4.0          # 弱倾斜斜率：回撤 12.5% → lean 触顶
LEAN_MAX = 0.50           # 弱势最多比 TWAP 快 50%
NO_LEAN_RULES = {1, 2, 6, 0, 7}   # ①慢清/②深套不砸底/⑥观察/⓪数据不足/⑦回前高(已是大额)：不前置

# 5.4 追踪止盈带：连续随基金自身波动缩放 trail = clip(1σ·√30d, 5%, 14%)
TRAIL_Z = 1.0
TRAIL_HORIZON_D = 30
TRAIL_FLOOR, TRAIL_CAP = 0.05, 0.14
TRAIL_DEFAULT = 0.08      # 波动样本不足时兜底带
# 5.6 趋势 band-hysteresis：现价在 MA20 ±TREND_BAND 内"贴线"不判方向，压日度 whipsaw
TREND_BAND = 0.01

MA_SHORT, MA_LONG = 20, 60
ROLL_HIGH_WIN = 250       # 滚动高水位窗口（深套 + 追踪止盈统一参照；次新基用全可用历史）
RSI_PERIOD = 14

# ── 持仓清单（去重 10 只，2026-05-31，中信证券 App）──────────────
# 持仓种子（首次生成 positions.csv 用；此后真值在 positions.csv，由前端"记录减仓"改写）
HOLDINGS_SEED = [
    # name, code, market_value, pnl_pct
    ("华夏卓越成长混合",     "024930", 152086.10,  0.52),
    ("睿远成长价值混合A",    "007119", 141325.43,  0.69),
    ("华夏红利价值混合",     "024915", 136386.96, -0.19),
    ("工银圆兴混合",         "009076", 133813.91,  0.12),
    ("中信保诚前瞻优势混合", "013610", 103306.75,  0.01),
    ("华安研究智选混合A",    "011692",  93525.56, -0.08),
    ("中欧医疗健康混合",     "003095",  50730.79, -0.56),
    ("朱雀产业智选混合",     "007880",  42118.97, -0.17),
    ("华夏兴阳一年持有混合", "009010",  32060.03, -0.54),
    ("泓德卓远混合A",        "010864",  22936.78, -0.11),
]
POSITIONS_CSV = OUTDIR / "positions.csv"


POS_COLS = ["code", "name", "market_value", "pnl_pct", "original_mv"]
ORIGINAL_SEED = {code: mv for name, code, mv, pnl in HOLDINGS_SEED}


def load_holdings():
    """读 positions.csv（持仓真值）；不存在则用种子生成。返回 (name,code,mv,pnl) 列表，mv>0。
    含 original_mv(不可变种子，累积减仓%加权用)；旧档缺该列则按 HOLDINGS_SEED 迁移补写。"""
    import csv as _csv
    if not POSITIONS_CSV.exists():
        with open(POSITIONS_CSV, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(POS_COLS)
            for name, code, mv, pnl in HOLDINGS_SEED:
                w.writerow([code, name, f"{mv:.2f}", pnl, f"{mv:.2f}"])
    with open(POSITIONS_CSV, encoding="utf-8") as f:
        rows = list(_csv.DictReader(f))
    changed = False
    for r in rows:
        if not r.get("original_mv"):   # 迁移：旧档无 original_mv → 用种子原值
            try:
                cur = float(r.get("market_value") or 0)
            except ValueError:
                cur = 0.0
            r["original_mv"] = f"{ORIGINAL_SEED.get(r['code'], cur):.2f}"
            changed = True
    if changed:
        with open(POSITIONS_CSV, "w", newline="", encoding="utf-8") as f:
            w = _csv.DictWriter(f, fieldnames=POS_COLS, extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow({c: r.get(c, "") for c in POS_COLS})
    out = []
    for r in rows:
        try:
            mv = float(r["market_value"])
        except (TypeError, ValueError):
            continue
        if mv > 0:
            pnl = float(r["pnl_pct"]) if r.get("pnl_pct") not in (None, "") else 0.0
            out.append((r["name"], r["code"], mv, pnl))
    return out
# 009010 华夏兴阳一年持有: 2021-01-04 申购 → ~2022-01 满1年持有期 → 已解锁, 现可自由赎回。
# 原"锁定"警告已移除(假约束)。若未来近12个月内有新增申购批次, 该批会重新锁1年, 届时再加回。
LOCKED = {}

ACTION_TEXT = {
    1: "①上升趋势：延长慢清(慢速TWAP)，让赢家多跑",
    7: "⑦回前高：反弹至前高附近，减仓兑现(清仓窗口)",
    8: "⑧回升未稳：破势后重回上升(<15日)，不授慢池，快池TWAP",
    2: "②深套：不加码砸底，仅走周 TWAP 逐步出",
    3: "③确认下降趋势：周 TWAP + 弱倾斜前置",
    4: "④震荡·破追踪带：周 TWAP + 弱倾斜前置",
    5: "⑤震荡·破MA20：周 TWAP + 轻前置",
    6: "⑥持有观察：仅走周 TWAP",
    0: "数据不足：仅走周 TWAP，指标待补",
}


# ── 工具函数 ─────────────────────────────────────────────────────
def fetch_nav(code: str, retries: int = 3, pause: float = 0.8) -> pd.DataFrame:
    """拉累计净值走势，retry + 升序。失败抛异常由上层隔离。"""
    last = None
    for i in range(retries):
        try:
            df = ak.fund_open_fund_info_em(symbol=code, indicator="累计净值走势")
            df = df.rename(columns={"净值日期": "date", "累计净值": "nav"})
            df["date"] = pd.to_datetime(df["date"])
            df["nav"] = pd.to_numeric(df["nav"], errors="coerce")
            df = df.dropna(subset=["nav"]).sort_values("date").reset_index(drop=True)
            if len(df) == 0:
                raise ValueError("空返回")
            return df
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(pause * (i + 1))
    raise last


def rsi(nav: pd.Series, period: int = 14) -> float:
    if len(nav) <= period:
        return float("nan")
    delta = nav.diff().dropna()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False).mean().iloc[-1]
    avg_loss = loss.ewm(alpha=1 / period, adjust=False).mean().iloc[-1]
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return float(100 - 100 / (1 + rs))


def _trend_label(latest, ma20, ma60) -> str:
    """趋势三态 + band-hysteresis（现价在 MA20 ±TREND_BAND 内=贴线，不判方向）。"""
    if pd.isna(ma20) or pd.isna(ma60):
        return "数据不足"
    if latest > ma20 * (1 + TREND_BAND) and ma20 > ma60:
        return "上升"
    if latest < ma20 * (1 - TREND_BAND) and ma20 < ma60:
        return "下降"
    return "震荡"


def fetch_unit_nav(code: str, retries: int = 2, pause: float = 0.6):
    """最新单位净值(份额估值用：金额 = 份额 × 单位净值)。失败返回 None。"""
    for i in range(retries):
        try:
            df = ak.fund_open_fund_info_em(symbol=code, indicator="单位净值走势")
            df["净值日期"] = pd.to_datetime(df["净值日期"])
            v = pd.to_numeric(df.sort_values("净值日期")["单位净值"], errors="coerce").dropna()
            if len(v):
                return float(v.iloc[-1])
        except Exception:  # noqa: BLE001
            time.sleep(pause * (i + 1))
    return None


def fetch_market_regime() -> str:
    """沪深300 现价 vs MA60 → 上下文提示（不进决策, per 5.5）。"""
    try:
        idx = ak.stock_zh_index_daily(symbol="sh000300")
        idx = idx.sort_values("date").reset_index(drop=True)
        close = pd.to_numeric(idx["close"], errors="coerce").dropna()
        ma60 = close.tail(60).mean()
        last = close.iloc[-1]
        pct = last / ma60 - 1
        state = "强(>MA60)" if last > ma60 else "弱(<MA60)"
        return f"沪深300 {state} 距MA60 {pct:+.1%}"
    except Exception as e:  # noqa: BLE001
        return f"市场regime取数失败({type(e).__name__})"


def fetch_fee(code: str) -> str:
    """尽力取赎回费率，取不到留 App 核对（per 5.7）。"""
    try:
        df = ak.fund_individual_detail_info_xq(symbol=code)
        # 雪球详情含赎回费阶梯；结构不稳定 → 只做存在性提示
        txt = df.to_string()
        if "赎回" in txt:
            return "见App(雪球有费率, 结构不稳)"
        return "见App核对"
    except Exception:
        return "见App核对"


# ── 指标 + 规则引擎 ──────────────────────────────────────────────
@dataclass
class FundResult:
    name: str
    code: str
    mv: float
    pnl_pct: float
    ok: bool = True
    err: str = ""
    latest: float = float("nan")
    latest_date: str = ""
    n: int = 0
    ma20: float = float("nan")
    ma60: float = float("nan")
    roll_high: float = float("nan")
    trail: float = float("nan")
    trail_trigger: float = float("nan")
    rsi: float = float("nan")
    ann_vol: float = float("nan")
    dist_ma20: float = float("nan")
    dist_ma60: float = float("nan")
    dist_roll: float = float("nan")      # 距滚动高（A6 统一参照）
    day_chg: float = float("nan")        # 最近一日涨跌% (累计净值环比)
    unit_nav: float = float("nan")       # 最新单位净值 (份额→金额换算用)
    trend: str = ""
    rule: int = 6
    action: str = ""
    base: float = 0.0                    # 强制保底量
    extra: float = 0.0                   # 信号额外量（可被超卖否决）
    clip: float = 0.0                    # = base + extra
    oversold: bool = False               # RSI<RSI_OVERSOLD（A1：仅此，不含创新低）
    short_hist: bool = False
    df: pd.DataFrame = field(default=None, repr=False)


def weeks_remaining_at(d, deadline) -> int:
    """从某日到给定 deadline 的剩余周数(向上取整，最少 1)。"""
    days_left = (deadline - d).days
    return max(1, math.ceil(days_left / 7)) if days_left > 0 else 1


def signal_at(nav, weeks_fast, weeks_slow) -> dict:
    """从累计净值序列(升序，截至某日)计算该日信号。analyze 与历史回填共用，确保口径一致。
    base 用 weeks_slow(上升①·延长清仓) 或 weeks_fast(其余·6周加速)。"""
    n = len(nav)
    latest = float(nav.iloc[-1])
    ma20 = float(nav.tail(MA_SHORT).mean()) if n >= MA_SHORT else float("nan")
    ma60 = float(nav.tail(MA_LONG).mean()) if n >= MA_LONG else float("nan")
    roll_high = float(nav.tail(min(ROLL_HIGH_WIN, n)).max())
    rsi_v = rsi(nav, RSI_PERIOD)
    rets = nav.pct_change().dropna()
    sigma_d = float(rets.tail(60).std()) if len(rets) >= 20 else float("nan")
    ann_vol = sigma_d * np.sqrt(252) if sigma_d == sigma_d else float("nan")
    trail = (float(np.clip(TRAIL_Z * sigma_d * np.sqrt(TRAIL_HORIZON_D), TRAIL_FLOOR, TRAIL_CAP))
             if sigma_d == sigma_d else TRAIL_DEFAULT)
    trail_trigger = roll_high * (1 - trail)
    dist_ma20 = latest / ma20 - 1 if ma20 == ma20 else float("nan")
    dist_ma60 = latest / ma60 - 1 if ma60 == ma60 else float("nan")
    dist_roll = latest / roll_high - 1 if roll_high == roll_high else float("nan")
    day_chg = float(nav.iloc[-1] / nav.iloc[-2] - 1) if n >= 2 else float("nan")
    trend = _trend_label(latest, ma20, ma60)
    oversold = (rsi_v == rsi_v and rsi_v < RSI_OVERSOLD)
    # ⑦回前高前置条件: 近 REC_LOOKBACK 日内曾非上升(破过势)。持续上升的真赢家不触发。
    ma20s = nav.rolling(MA_SHORT).mean()
    ma60s = nav.rolling(MA_LONG).mean()
    recent_break = any(
        _trend_label(float(nav.iloc[i]), float(ma20s.iloc[i]), float(ma60s.iloc[i])) != "上升"
        for i in range(max(0, n - REC_LOOKBACK), n)
    )
    # 规则引擎（优先级）；深套/回前高用 dist_roll（统一滚动高，A6）
    if dist_roll == dist_roll and dist_roll >= -REC_BAND and recent_break:
        rule = 7   # 回前高：反弹至前高附近就卖，优先于①（防反弹强→重回上升→反而转慢池）
    elif trend == "上升" and not recent_break:
        rule = 1   # 慢池资格 = 持续上升(近15日无破势)的真赢家
    elif trend == "上升":
        rule = 8   # 回升未稳：破势后重回上升(<15日) → 不授慢池(修非单调: 半反弹曾掉到12.5%<地板)
    elif trend == "下降" and dist_roll <= -DEEP:
        rule = 2
    elif trend == "下降":
        rule = 3
    elif latest <= trail_trigger:
        rule = 4
    elif trend == "震荡" and ma20 > ma60 and latest < ma20:
        rule = 5
    elif trend == "数据不足":
        rule = 0
    else:
        rule = 6
    # 周频截止日 TWAP：base = 1/剩余周数；extra = 弱倾斜前置（超卖只挡 extra）
    # ①用慢速(延长 deadline)；⑦回前高固定大额；其余用快速(软化下限见 weeks_fast 调用侧)
    if rule == 7:
        base = REC_CLIP
    else:
        base = min(1.0, 1.0 / (weeks_slow if rule == 1 else weeks_fast))
    depth = max(0.0, -dist_roll) if dist_roll == dist_roll else 0.0
    lean = 0.0 if rule in NO_LEAN_RULES else min(LEAN_SLOPE * depth, LEAN_MAX)
    extra_vetoed = oversold and lean > 0
    if extra_vetoed:
        lean = 0.0
    extra = base * lean
    return dict(n=n, latest=latest, short_hist=n < MA_LONG, ma20=ma20, ma60=ma60,
                roll_high=roll_high, rsi=rsi_v, ann_vol=ann_vol, trail=trail,
                trail_trigger=trail_trigger, dist_ma20=dist_ma20, dist_ma60=dist_ma60,
                dist_roll=dist_roll, day_chg=day_chg, trend=trend, oversold=oversold,
                rule=rule, base=base, extra=extra, clip=min(base + extra, 1.0),
                extra_vetoed=extra_vetoed)


def analyze(name, code, mv, pnl_pct, weeks_fast, weeks_slow) -> FundResult:
    r = FundResult(name=name, code=code, mv=mv, pnl_pct=pnl_pct)
    try:
        df = fetch_nav(code)
    except Exception as e:  # noqa: BLE001
        r.ok = False
        r.err = f"{type(e).__name__}: {e}"
        return r

    r.df = df
    r.latest_date = df["date"].iloc[-1].strftime("%Y-%m-%d")
    # 单位净值(份额换算用)。取数失败不回退累计净值——分红基两者不同, 静默错值比缺值更糟;
    # 缺值时前端拒绝换算并提示, 下次扫描自愈。
    u = fetch_unit_nav(code)
    r.unit_nav = u if u else float("nan")
    s = signal_at(df["nav"], weeks_fast, weeks_slow)
    for k in ("n", "latest", "short_hist", "ma20", "ma60", "roll_high", "rsi", "ann_vol",
              "trail", "trail_trigger", "dist_ma20", "dist_ma60", "dist_roll", "day_chg",
              "trend", "oversold", "rule", "base", "extra", "clip"):
        setattr(r, k, s[k])
    r.action = ACTION_TEXT[r.rule]
    if s["extra_vetoed"]:
        r.action += "｜超卖否决前置(RSI<{}，仅走 TWAP)".format(RSI_OVERSOLD)
    return r


# ── 主流程 ───────────────────────────────────────────────────────
def main():
    # 周频截止日 TWAP：按真实今日到 deadline 的剩余周数定步
    today = datetime.now().date()
    weeks_fast = max(weeks_remaining_at(today, DEADLINE_FAST), SOFT_FLOOR_WEEKS)  # 软化: 不悬崖到 100%/周
    if today > DEADLINE_SLOW:   # 硬外沿: 软化有边界——8-31 后全书回到"剩1周清光"(thesis 到期)
        weeks_fast = 1
    weeks_slow = weeks_remaining_at(today, DEADLINE_SLOW)
    twap_fast = min(1.0, 1.0 / weeks_fast)
    twap_slow = min(1.0, 1.0 / weeks_slow)

    print("=" * 90)
    print(f"基金清仓信号 v5(双速TWAP) | 快 {weeks_fast}周→{DEADLINE_FAST} {twap_fast:.1%}/只 | "
          f"上升慢 {weeks_slow}周→{DEADLINE_SLOW} {twap_slow:.1%}/只")
    print("=" * 90)

    regime = fetch_market_regime()
    print(f"市场 regime（仅提示, 不进决策）: {regime}\n")

    holdings = load_holdings()
    results = []
    for name, code, mv, pnl in holdings:
        print(f"  拉取 {code} {name} ...", end=" ")
        try:
            r = analyze(name, code, mv, pnl, weeks_fast, weeks_slow)
            if r.ok:
                print(f"OK n={r.n} {r.trend} -> 规则{r.rule} 本周clip={r.clip:.1%}")
            else:
                print(f"失败跳过: {r.err}")
        except Exception as e:  # noqa: BLE001  最后兜底, 绝不中断
            traceback.print_exc()
            r = FundResult(name=name, code=code, mv=mv, pnl_pct=pnl,
                           ok=False, err=f"{type(e).__name__}: {e}")
            print(f"异常跳过: {r.err}")
        results.append(r)
        time.sleep(0.4)

    ok = [r for r in results if r.ok]
    failed = [r for r in results if not r.ok]
    data_date = max((r.latest_date for r in ok), default=today.isoformat())

    # ── 两个清仓池：上升①(慢速·延长) vs 其余(6周·加速) ──
    slow_mv = sum(r.mv for r in ok if r.rule == 1)
    fast_mv = sum(r.mv for r in ok if r.rule != 1)
    suggested_total = sum(r.mv * r.clip for r in ok)
    cad = {"weeks_fast": weeks_fast, "weeks_slow": weeks_slow,
           "twap_fast": twap_fast, "twap_slow": twap_slow,
           "fast_mv": fast_mv, "slow_mv": slow_mv, "suggested_total": suggested_total,
           "deadline_fast": DEADLINE_FAST.isoformat(), "deadline_slow": DEADLINE_SLOW.isoformat()}

    # ── 汇总表 ──
    rows = []
    for r in results:
        if not r.ok:
            rows.append({
                "基金名称": r.name, "代码": r.code, "市值": r.mv,
                "建议动作": f"❌取数失败: {r.err}", "数据日": "", "最新净值": np.nan,
                "MA20": np.nan, "MA60": np.nan, "滚动最高": np.nan, "追踪带%": np.nan,
                "距MA20%": np.nan, "距滚动高%": np.nan, "趋势": "", "RSI": np.nan,
                "波动(年化)": np.nan, "保底%": np.nan, "额外%": np.nan,
                "建议卖出%": np.nan, "建议卖出¥": np.nan,
                "vsTWAP%": np.nan, "费率": "见App核对", "锁定/备注": LOCKED.get(r.code, ""),
            })
            continue
        dev = r.clip - (twap_slow if r.rule == 1 else twap_fast)   # vs 本池 TWAP
        rows.append({
            "基金名称": r.name, "代码": r.code, "市值": round(r.mv, 2),
            "建议动作": r.action, "数据日": r.latest_date, "最新净值": round(r.latest, 4),
            "MA20": round(r.ma20, 4) if r.ma20 == r.ma20 else np.nan,
            "MA60": round(r.ma60, 4) if r.ma60 == r.ma60 else np.nan,
            "滚动最高": round(r.roll_high, 4), "追踪带%": r.trail,
            "距MA20%": r.dist_ma20, "距滚动高%": r.dist_roll, "趋势": r.trend,
            "RSI": round(r.rsi, 1) if r.rsi == r.rsi else np.nan,
            "波动(年化)": r.ann_vol, "保底%": r.base, "额外%": r.extra,
            "建议卖出%": r.clip,
            "建议卖出¥": round(r.mv * r.clip, 0), "vsTWAP%": dev,
            "费率": "见App核对",   # 原 fetch_fee 每次扫描打 akshare 且返回恒为占位 → 移除网络调用
            "锁定/备注": (LOCKED.get(r.code, "") + (" 次新基样本不足" if r.short_hist else "")).strip(),
        })

    df = pd.DataFrame(rows)
    # 排序：强势(rule1)沉底，其余按建议卖出%降序（卖出队列）
    rule_map = {r.code: r.rule for r in results}
    df["_rule"] = df["代码"].map(rule_map).fillna(99)
    df["_clip"] = df["建议卖出%"].fillna(-1)
    df = df.sort_values(["_clip", "市值"], ascending=[False, False]).drop(columns=["_rule", "_clip"])

    # ── 输出 Excel ──
    xlsx = OUTDIR / "基金技术出场信号.xlsx"
    write_excel(df, xlsx, regime, data_date, cad)
    print(f"\n写出 {xlsx}")

    # ── 输出 JSON（前端数据契约, schema 见 task/fund_exit_FE_handoff.md）──
    write_json(results, OUTDIR / "fund_signals.json", regime, data_date, cad)
    print(f"写出 {OUTDIR / 'fund_signals.json'}")

    # ── 每日信号日志（按 净值日×代码 去重 upsert）──
    write_signal_log(results, OUTDIR / "signal_log.csv")

    # ── 行业透视（display-only; 文件新鲜则跳过, 失败不影响主扫描）──
    try:
        from fund_exit.fund_sectors import refresh_if_stale
    except ImportError:
        from fund_sectors import refresh_if_stale
    refresh_if_stale(holdings)

    # ── 图 ──
    chart_dir = OUTDIR / "charts"
    chart_dir.mkdir(exist_ok=True)
    for r in ok:
        try:
            plot_fund(r, chart_dir)
        except Exception as e:  # noqa: BLE001
            print(f"  绘图失败 {r.code}: {e}")
    print(f"写出 {len(ok)} 张图 -> {chart_dir}")

    # ── 文字小结 ──
    print_summary(ok, failed, regime, cad)


def write_excel(df, path, regime, data_date, cad):
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="信号", startrow=4)
        ws = xw.sheets["信号"]
        ws["A1"] = f"基金清仓信号(双速TWAP)  |  数据日 {data_date}  |  纪律工具, 非投资建议"
        ws["A2"] = f"市场regime(仅提示): {regime}"
        ws["A3"] = (f"加速(非上升) ¥{cad['fast_mv']:,.0f} · 剩 {cad['weeks_fast']}周 · {cad['twap_fast']:.1%}/只  |  "
                    f"延长(上升①) ¥{cad['slow_mv']:,.0f} · 剩 {cad['weeks_slow']}周 · {cad['twap_slow']:.1%}/只  |  "
                    f"本周建议合计 ¥{cad['suggested_total']:,.0f}")
        ws["A1"].font = Font(bold=True, size=12)

        # 百分比格式列
        pct_cols = {"追踪带%", "距MA20%", "距滚动高%", "波动(年化)", "保底%", "额外%", "建议卖出%", "vsTWAP%"}
        money_cols = {"市值", "建议卖出¥"}
        headers = {c.value: c.column_letter for c in ws[5]}
        for col, letter in headers.items():
            if col in pct_cols:
                for cell in ws[letter][5:]:
                    cell.number_format = "0.0%"
            elif col in money_cols:
                for cell in ws[letter][5:]:
                    cell.number_format = "#,##0"
            elif col in {"最新净值", "MA20", "MA60", "滚动最高"}:
                for cell in ws[letter][5:]:
                    cell.number_format = "0.0000"
        for cell in ws[5]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")
        widths = {"基金名称": 22, "代码": 9, "建议动作": 42, "费率": 22, "锁定/备注": 28}
        for col, letter in headers.items():
            ws.column_dimensions[letter].width = widths.get(col, 12)
        ws.freeze_panes = "A6"


def _num(x):
    """NaN/inf -> None（JSON 友好）。"""
    try:
        f = float(x)
        return None if (f != f or f in (float("inf"), float("-inf"))) else round(f, 6)
    except (TypeError, ValueError):
        return None


def write_json(results, path, regime, data_date, cad):
    """前端数据契约。schema 见 task/fund_exit_FE_handoff.md。"""
    import json
    from datetime import datetime

    ok = [r for r in results if r.ok]
    total_mv = sum(r.mv for r in results)

    funds = []
    for r in results:
        funds.append({
            "name": r.name, "code": r.code, "mv": _num(r.mv), "pnl_pct": _num(r.pnl_pct),
            "ok": r.ok, "err": r.err,
            "latest": _num(r.latest), "latest_date": r.latest_date, "n": r.n,
            "short_hist": r.short_hist,
            "ma20": _num(r.ma20), "ma60": _num(r.ma60),
            "roll_high": _num(r.roll_high), "trail": _num(r.trail),
            "trail_trigger": _num(r.trail_trigger),
            "rsi": _num(r.rsi), "ann_vol": _num(r.ann_vol),
            "dist_ma20": _num(r.dist_ma20), "dist_ma60": _num(r.dist_ma60),
            "dist_roll": _num(r.dist_roll), "day_chg": _num(r.day_chg),
            "unit_nav": _num(r.unit_nav),
            "trend": r.trend, "rule": r.rule, "action": r.action,
            "base": _num(r.base), "extra": _num(r.extra),
            "clip": _num(r.clip), "clip_amt": _num(r.mv * r.clip) if r.ok else None,
            "vs_twap": _num(r.clip - (cad["twap_slow"] if r.rule == 1 else cad["twap_fast"])) if r.ok else None,
            "oversold": r.oversold,
            "locked": LOCKED.get(r.code, ""),
            "chart": f"charts/{r.code}_{r.name}.png" if r.ok else None,
        })

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "data_date": data_date,
        "market_regime": regime,
        "disclaimer": "纪律工具，非投资建议；阈值先验非优化；费率/锁定/赎回以中信App为准。",
        "cadence": {                       # 双速周频 TWAP
            "fast": {"deadline": cad["deadline_fast"], "weeks": cad["weeks_fast"],
                     "twap": _num(cad["twap_fast"]), "mv": _num(cad["fast_mv"])},
            "slow": {"deadline": cad["deadline_slow"], "weeks": cad["weeks_slow"],
                     "twap": _num(cad["twap_slow"]), "mv": _num(cad["slow_mv"])},
            "op_freq": "每周一次(美东晚间下单, 成交落次日中国净值)",
        },
        "params": {"DEEP": DEEP, "rsi_oversold": RSI_OVERSOLD,
                   "trail_formula": f"clip({TRAIL_Z}σ·√{TRAIL_HORIZON_D}d, {TRAIL_FLOOR}, {TRAIL_CAP})",
                   "trend_band": TREND_BAND,
                   "model": "clip = base(周频TWAP: ①慢速延长/⑦回前高固定{:.0%}/其余快速·软化下限{}周) + extra(弱倾斜前置, 可被RSI<{}否决)".format(REC_CLIP, SOFT_FLOOR_WEEKS, RSI_OVERSOLD),
                   "recovery": {"band": REC_BAND, "clip": REC_CLIP, "lookback": REC_LOOKBACK},
                   "lean": f"min({LEAN_SLOPE}×回撤, {LEAN_MAX})×base"},
        "account": {
            "total_mv": _num(total_mv),
            "fast_mv": _num(cad["fast_mv"]), "slow_mv": _num(cad["slow_mv"]),
            "slow_pct": _num(cad["slow_mv"] / total_mv) if total_mv else 0,
            "suggested_total": _num(cad["suggested_total"]),
        },
        "funds": funds,
    }
    # 原子写：先写临时文件再 rename，避免日度扫描重写时被前端读到半截
    import os
    tmp = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


_RULE_NAME = {1: "①上升持有", 7: "⑦回前高", 8: "⑧回升未稳", 2: "②深套", 3: "③确认下降", 4: "④破追踪带",
              5: "⑤破MA20", 6: "⑥观察", 0: "数据不足"}


def write_signal_log(results, path):
    """每只基金、每个净值日的信号分类(rule)+ 指标，按 (date,code) upsert。
    3 次/日扫描幂等：同一净值日重跑只覆盖该日该只这一行。"""
    import csv as _csv
    cols = ["date", "code", "name", "rule", "rule_name", "trend", "clip",
            "base", "extra", "dist_roll", "day_chg", "rsi", "oversold", "mv", "recorded_at"]
    now = datetime.now().isoformat(timespec="seconds")
    new = {}
    for r in results:
        if not r.ok or not r.latest_date:
            continue
        new[(r.latest_date, r.code)] = {
            "date": r.latest_date, "code": r.code, "name": r.name,
            "rule": r.rule, "rule_name": _RULE_NAME.get(r.rule, ""), "trend": r.trend,
            "clip": _num(r.clip), "base": _num(r.base), "extra": _num(r.extra),
            "dist_roll": _num(r.dist_roll), "day_chg": _num(r.day_chg),
            "rsi": round(r.rsi, 1) if r.rsi == r.rsi else "",
            "oversold": int(r.oversold), "mv": _num(r.mv), "recorded_at": now,
        }
    merged = {}
    if path.exists():
        with open(path, encoding="utf-8") as f:
            for row in _csv.DictReader(f):
                merged[(row.get("date"), row.get("code"))] = row
    merged.update(new)   # 同 (date,code) 以本次为准
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for k in sorted(merged, key=lambda x: (x[0] or "", x[1] or "")):
            w.writerow({c: merged[k].get(c, "") for c in cols})


def backfill_signal_log(days=60):
    """回放近 N 个交易日的信号，一次性补进 signal_log.csv（幂等 upsert，口径同 signal_at）。
    mv 留空（历史持仓未知）；rule/趋势/距滚高/RSI/日涨跌 精确重建。"""
    import csv as _csv
    path = OUTDIR / "signal_log.csv"
    cols = ["date", "code", "name", "rule", "rule_name", "trend", "clip", "base", "extra",
            "dist_roll", "day_chg", "rsi", "oversold", "mv", "recorded_at"]
    now = datetime.now().isoformat(timespec="seconds") + " backfill"
    merged = {}
    if path.exists():
        with open(path, encoding="utf-8") as f:
            for row in _csv.DictReader(f):
                merged[(row.get("date"), row.get("code"))] = row
    print(f"回放近 {days} 交易日信号 …")
    for name, code, mv, pnl in load_holdings():
        print(f"  {code} {name} ...", end=" ")
        try:
            df = fetch_nav(code)
        except Exception as e:  # noqa: BLE001
            print(f"失败: {e}")
            continue
        nav, dates, n = df["nav"], df["date"], len(df)
        start = max(MA_LONG, n - days)   # 需 ≥MA_LONG 历史才有趋势
        cnt = 0
        for i in range(start, n):
            d = dates.iloc[i].strftime("%Y-%m-%d")
            d_i = dates.iloc[i].date()
            wf = max(weeks_remaining_at(d_i, DEADLINE_FAST), SOFT_FLOOR_WEEKS)
            if d_i > DEADLINE_SLOW:
                wf = 1
            s = signal_at(nav.iloc[:i + 1], wf, weeks_remaining_at(d_i, DEADLINE_SLOW))
            merged[(d, code)] = {
                "date": d, "code": code, "name": name,
                "rule": s["rule"], "rule_name": _RULE_NAME.get(s["rule"], ""), "trend": s["trend"],
                "clip": _num(s["clip"]), "base": _num(s["base"]), "extra": _num(s["extra"]),
                "dist_roll": _num(s["dist_roll"]), "day_chg": _num(s["day_chg"]),
                "rsi": round(s["rsi"], 1) if s["rsi"] == s["rsi"] else "",
                "oversold": int(s["oversold"]), "mv": "", "recorded_at": now,
            }
            cnt += 1
        print(f"{cnt} 天")
        time.sleep(0.4)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for k in sorted(merged, key=lambda x: (x[0] or "", x[1] or "")):
            w.writerow({c: merged[k].get(c, "") for c in cols})
    print(f"signal_log 现有 {len(merged)} 行")


def plot_fund(r: FundResult, chart_dir: Path):
    df = r.df.tail(min(ROLL_HIGH_WIN, r.n)).copy()
    nav = df["nav"]
    dates = df["date"]
    fig, ax = plt.subplots(figsize=(10, 4.5))
    ax.plot(dates, nav, label="累计净值", color="#1f3a5f", lw=1.4)
    if r.n >= MA_SHORT:
        ax.plot(dates, nav.rolling(MA_SHORT).mean(), label=f"MA{MA_SHORT}", color="#e08a00", lw=1.0)
    if r.n >= MA_LONG:
        ax.plot(dates, nav.rolling(MA_LONG).mean(), label=f"MA{MA_LONG}", color="#2e7d32", lw=1.0)
    # 滚动高水位点（深套/追踪的统一参照，A6）
    hi_idx = df["nav"].idxmax()
    ax.scatter(df.loc[hi_idx, "date"], df.loc[hi_idx, "nav"], color="red", zorder=5, s=30, label="滚动高")
    # 追踪止盈触发位
    ax.axhline(r.trail_trigger, color="purple", ls="--", lw=1.0,
               label=f"追踪止盈触发 {r.trail_trigger:.4f} (滚动高-{r.trail*100:.1f}% σ缩放)")
    ax.set_title(f"{r.name} ({r.code})  {r.trend}  -> {r.action.split('｜')[0]}", fontsize=11)
    ax.legend(fontsize=8, loc="best")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(chart_dir / f"{r.code}_{r.name}.png", dpi=110)
    plt.close(fig)


def print_summary(ok, failed, regime, cad):
    print("\n" + "=" * 90)
    print("文字小结")
    print("=" * 90)
    by_rule = {}
    for r in ok:
        by_rule.setdefault(r.rule, []).append(r)

    sell = [r for r in ok if r.clip > 0]
    uptrend = by_rule.get(1, [])

    print(f"\n市场: {regime}")
    print(f"加速(非上升) ¥{cad['fast_mv']:,.0f}·剩{cad['weeks_fast']}周·{cad['twap_fast']:.1%}/只 | "
          f"延长(上升①) ¥{cad['slow_mv']:,.0f}·剩{cad['weeks_slow']}周·{cad['twap_slow']:.1%}/只 | "
          f"本周建议合计 ¥{cad['suggested_total']:,.0f}")

    print(f"\n【本周建议卖出 {len(sell)} 只】(clip = 周TWAP base + 弱倾斜 extra)")
    for r in sorted(sell, key=lambda x: -x.clip):
        pool = "慢" if r.rule == 1 else "快"
        print(f"  {r.code} {r.name:<18} [{pool}]{r.action.split('｜')[0]:<22} "
              f"卖{r.clip:.1%}(TWAP{r.base:.1%}+前置{r.extra:.1%}) ≈¥{r.mv*r.clip:,.0f}  "
              f"距滚高{r.dist_roll:+.1%} RSI{r.rsi:.0f}{' 超卖' if r.oversold else ''}")

    print(f"\n【上升趋势·延长清仓 {len(uptrend)} 只】(慢速 TWAP {cad['twap_slow']:.1%}/周, 给赢家更多时间)")
    for r in uptrend:
        print(f"  {r.code} {r.name:<18} 距滚高{r.dist_roll:+.1%} RSI{r.rsi:.0f}  慢卖{r.clip:.1%}/周 ≈¥{r.mv*r.clip:,.0f}")

    # 深套两只专项
    print("\n【深套专项】中欧医疗 003095 / 华夏兴阳 009010")
    for code in ("003095", "009010"):
        r = next((x for x in ok if x.code == code), None)
        if r is None:
            print(f"  {code}: 取数失败, 见上方报告")
            continue
        note = LOCKED.get(code, "")
        print(f"  {code} {r.name}: {r.trend} 距滚高{r.dist_roll:+.1%} 距MA20{r.dist_ma20:+.1%} "
              f"-> {r.action.split('｜')[0]} (卖{r.clip:.0%})")
        if r.rule == 2:
            print(f"      深套: 不前置, 仅走周 TWAP {r.base:.1%}/周逐步出(避开 capitulation 加码)")
        if note:
            print(f"      {note}")

    if failed:
        print(f"\n【取数失败 {len(failed)} 只】(已跳过, 不影响其余)")
        for r in failed:
            print(f"  {r.code} {r.name}: {r.err}")

    total_mv = sum(r.mv for r in ok) + sum(r.mv for r in failed)
    slow_pct = cad["slow_mv"] / total_mv if total_mv else 0
    print(f"\n注: 双速清仓(PM 2026-06-04)—— 非上升仓 {cad['weeks_fast']}周加速清, 上升①仓 "
          f"¥{cad['slow_mv']:,.0f}(≈账户{slow_pct:.0%}) {cad['weeks_slow']}周延长慢清(给赢家更多时间, 仍逐步出); "
          "阈值先验非优化; 费率/锁定/具体赎回以中信App为准。")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "backfill":
        backfill_signal_log(int(sys.argv[2]) if len(sys.argv) > 2 else 60)
    else:
        main()
