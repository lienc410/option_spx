"""
web/partnership_book.py — read-only parser for the family Partnership Shares model.

Source of truth is the pure-formula Excel workbook
`research/book_management/Partnership_Shares_v3.5.xlsx` (20 sheets, ~1,290
formulas, zero formula errors). This module NEVER writes back to the workbook —
it only reads the *cached* computed values (openpyxl ``data_only=True``) from the
output sheets and exposes them to the dashboard.

Why parse instead of reimplement: re-deriving the 1,290 formulas in Python would
create a second source of truth and the exact doc-drift failure mode the PM has
flagged as catastrophic. The engine stays in Excel; the dashboard mirrors it.

Two public entry points:
  - ``read_book()``  → parsed Consolidation / SW_Summary / ET_Summary outputs.
  - ``live_nlv()``   → best-effort live broker NLV overlay vs latest reconciled
                       Excel snapshot (informational; never written back).
"""
from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

log = logging.getLogger(__name__)

# research/book_management/ lives next to web/ under the repo root.
_REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = _REPO_ROOT / "research" / "book_management" / "Partnership_Shares_v3.5.xlsx"

# --- Google Drive source (preferred when a service account is configured) ---
# The workbook is a native Google Sheet; we export it to xlsx and read the
# cached computed values (export preserves them). Source of truth stays the live
# Sheet — we only read, never write. Falls back to the local file on any failure.
#   PARTNERSHIP_DRIVE_FILE_ID    — Sheet id (defaults to the PM's known file)
#   GOOGLE_SERVICE_ACCOUNT_JSON  — path to the SA key json; absent → local-only
DEFAULT_DRIVE_FILE_ID = "1_RsGTprzu_5c56128n5AkJP4bIDbqF9OSmdef8CFqFU"
DRIVE_EXPORT_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# Member rows in Consolidation (A5:I9) and SW_Summary (A5:I9) are positional.
_CONSOLIDATION_MEMBER_ROWS = range(5, 10)   # Chinchaung, Lien, Xinzhong/CXZ, Alvin, Lucas
_CONSOLIDATION_TOTAL_ROW = 10
_CONSOLIDATION_AUM_ROWS = {"total": 13, "schwab": 15, "etrade": 16}

# SW_Summary "RETURNS BY YEAR" block (rows 23-25); columns C-G are per-partner $.
_SW_YEAR_ROWS = range(23, 26)
_SW_YEAR_PARTNER_COLS = {  # column letter → partner display name
    "C": "Chinchaung", "D": "Lien", "E": "Xinzhong/CXZ", "F": "Alvin", "G": "Lucas",
}
# SW_Summary per-member name → its key in the by_year "partners" dict.
_SW_YEAR_PARTNER_COLS_BY_SW = {
    "Chinchaung": "Chinchaung", "Lien": "Lien", "Xinzhong": "Xinzhong/CXZ",
    "Alvin": "Alvin", "Lucas": "Lucas",
}

# SW_Summary per-member rows (A5:I9): contrib/distrib/net/pnl/balance/share/simple/IRR.
_SW_MEMBER_ROWS = range(5, 10)

# ET_Summary partner rows (A5:G6).
_ET_SUMMARY_ROWS = range(5, 7)

# E*Trade per-member yearly returns. Pre-merge the two partners had SEPARATE
# accounts (Lien=3370+4511 Nov→Nov window, 2024 incomplete; CXZ=6842 calendar
# year from May'24), so there is NO shared-pool TWR like CCC-354. Rows hold
# (period, P&L col E, return col F) on each sheet.
_ET_LIEN_YEAR_ROWS = {"2024": 20, "2025": 21, "2026 YTD": 22}   # Lien_ETrade
_ET_CXZ_YEAR_ROWS = {"2024": 26, "2025": 27, "2026 YTD": 28}    # CXZ_ETrade

# Join Consolidation member → its SW_Summary name and ET_Summary name (None = no
# E*Trade account). Xinzhong (Schwab) and CXZ (E*Trade) are the same person.
_MEMBER_JOIN = {
    "Chinchaung":      {"sw": "Chinchaung", "et": None},
    "Lien":            {"sw": "Lien",       "et": "Lien"},
    "Xinzhong / CXZ":  {"sw": "Xinzhong",   "et": "CXZ"},
    "Alvin":           {"sw": "Alvin",      "et": None},
    "Lucas":           {"sw": "Lucas",      "et": None},
}

# --- module cache keyed on source version (re-parse only when the source changes) ---
#   key = ("drive", modifiedTime) | ("local", mtime)
_cache: dict[str, Any] = {"key": None, "payload": None}


def _f(value: Any) -> float | None:
    """Coerce a cell to float; tolerate None / strings."""
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell(ws, ref: str) -> Any:
    return ws[ref].value


def read_book(force: bool = False) -> dict[str, Any]:
    """Parse the workbook's output sheets into a JSON-able dict.

    Source precedence: Google Drive (when a service account is configured) →
    local file. Fail-soft everywhere: on any error returns
    ``{"available": False, "message": ...}`` so the page degrades gracefully.
    Cached by source version (Drive ``modifiedTime`` / local mtime) so repeated
    calls within the same workbook version are cheap.
    """
    if _drive_configured():
        payload = _read_from_drive(force)
        if payload is not None:
            return payload
        # Drive configured but failed → fall through to local copy if present.
        log.warning("partnership_book: Drive source failed, falling back to local file")
    return _read_from_local(force)


def _finalize(payload_or_none, *, source: str, source_detail: str, version_iso: str,
              cache_key) -> dict[str, Any]:
    """Wrap a parsed payload with provenance + cache it."""
    payload = payload_or_none
    payload["available"] = True
    payload["source"] = source              # "google_drive" | "local_file"
    payload["source_file"] = source_detail
    payload["workbook_mtime"] = version_iso
    _cache["key"] = cache_key
    _cache["payload"] = payload
    return payload


def _parse_bytes_or_path(src) -> dict[str, Any] | None:
    """Open (path or file-like) with openpyxl data_only and parse. None on error."""
    try:
        import openpyxl
    except ImportError:
        log.warning("partnership_book: openpyxl not installed")
        return None
    try:
        # Not read_only: the Drive xlsx export omits stored sheet dimensions,
        # which makes ws.max_row None in read_only mode. The workbook is ~130KB,
        # so a full (dimension-computing) load is cheap and works for both
        # Drive-exported and locally-saved copies.
        wb = openpyxl.load_workbook(src, data_only=True)
    except Exception:
        log.warning("partnership_book: failed to open workbook", exc_info=True)
        return None
    try:
        return _parse(wb)
    except Exception:
        log.warning("partnership_book: failed to parse workbook", exc_info=True)
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _read_from_local(force: bool = False) -> dict[str, Any]:
    if not WORKBOOK_PATH.exists():
        return {
            "available": False,
            "message": "账本文件未找到（Drive 未配置且本地 research/book_management/ 缺文件）",
        }
    mtime = WORKBOOK_PATH.stat().st_mtime
    key = ("local", mtime)
    if not force and _cache["key"] == key and _cache["payload"] is not None:
        return _cache["payload"]

    parsed = _parse_bytes_or_path(WORKBOOK_PATH)
    if parsed is None:
        return {"available": False, "message": "账本文件无法读取或解析（可能已损坏/表结构变动）"}

    return _finalize(
        parsed,
        source="local_file",
        source_detail=str(WORKBOOK_PATH.relative_to(_REPO_ROOT)),
        version_iso=datetime.fromtimestamp(mtime, tz=timezone.utc).isoformat(timespec="seconds"),
        cache_key=key,
    )


# ----------------------------- Google Drive source -----------------------------

def _drive_file_id() -> str:
    return os.getenv("PARTNERSHIP_DRIVE_FILE_ID", DEFAULT_DRIVE_FILE_ID)


def _drive_sa_key_path() -> str | None:
    return os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")


def _drive_configured() -> bool:
    """Drive is the source only when a service account key path is set and exists."""
    key = _drive_sa_key_path()
    return bool(key) and Path(key).expanduser().is_file()


def _drive_service():
    """Build a read-only Drive v3 service from the service account key. None on error."""
    try:
        from google.oauth2 import service_account  # type: ignore
        from googleapiclient.discovery import build  # type: ignore
    except Exception:
        log.warning("partnership_book: google client libs not installed")
        return None
    try:
        creds = service_account.Credentials.from_service_account_file(
            str(Path(_drive_sa_key_path()).expanduser()), scopes=DRIVE_SCOPES
        )
        return build("drive", "v3", credentials=creds, cache_discovery=False)
    except Exception:
        log.warning("partnership_book: failed to build Drive service", exc_info=True)
        return None


def _read_from_drive(force: bool = False) -> dict[str, Any] | None:
    """Export the Sheet → xlsx → parse. Returns payload, or None to signal the
    caller to fall back to the local file. A successful but *unavailable* parse
    (e.g. permission error) returns a fail-soft dict rather than None only when
    there is no local fallback — but to keep behaviour predictable we return None
    on any fetch error so local can take over."""
    service = _drive_service()
    if service is None:
        return None

    file_id = _drive_file_id()

    # Cheap metadata check first: skip re-export if modifiedTime is unchanged.
    try:
        meta = service.files().get(
            fileId=file_id, fields="modifiedTime,name", supportsAllDrives=True
        ).execute()
    except Exception:
        log.warning("partnership_book: Drive metadata fetch failed", exc_info=True)
        return None

    modified = meta.get("modifiedTime") or ""
    key = ("drive", modified)
    if not force and _cache["key"] == key and _cache["payload"] is not None:
        return _cache["payload"]

    try:
        data = service.files().export(fileId=file_id, mimeType=DRIVE_EXPORT_MIME).execute()
    except Exception:
        log.warning("partnership_book: Drive export failed", exc_info=True)
        return None
    if not data:
        return None

    parsed = _parse_bytes_or_path(io.BytesIO(data))
    if parsed is None:
        return None

    return _finalize(
        parsed,
        source="google_drive",
        source_detail=f"Google Drive · {meta.get('name', 'Partnership Shares')}",
        version_iso=modified,
        cache_key=key,
    )


def _parse(wb) -> dict[str, Any]:
    cons = wb["Consolidation"]
    swsum = wb["SW_Summary"]
    etsum = wb["ET_Summary"]
    swsnap = wb["SW_Snapshots"]
    etsnap = wb["ET_Snapshots"]
    cxz_et = wb["CXZ_ETrade"]
    lien_et = wb["Lien_ETrade"]

    # --- Consolidation: per-member cross-broker rollup (A5:I9) + total (row 10) ---
    members: list[dict[str, Any]] = []
    for r in _CONSOLIDATION_MEMBER_ROWS:
        name = _cell(cons, f"A{r}")
        if not name:
            continue
        members.append({
            "name": str(name),
            "schwab_value": _f(_cell(cons, f"B{r}")),
            "etrade_value": _f(_cell(cons, f"C{r}")),
            "total_value": _f(_cell(cons, f"D{r}")),
            "schwab_pct": _f(_cell(cons, f"E{r}")),
            "etrade_pct": _f(_cell(cons, f"F{r}")),
            "contrib": _f(_cell(cons, f"G{r}")),
            "pnl": _f(_cell(cons, f"H{r}")),
            "return_pct": _f(_cell(cons, f"I{r}")),
        })

    tr = _CONSOLIDATION_TOTAL_ROW
    total = {
        "schwab_value": _f(_cell(cons, f"B{tr}")),
        "etrade_value": _f(_cell(cons, f"C{tr}")),
        "total_value": _f(_cell(cons, f"D{tr}")),
        "schwab_pct": _f(_cell(cons, f"E{tr}")),
        "etrade_pct": _f(_cell(cons, f"F{tr}")),
        "contrib": _f(_cell(cons, f"G{tr}")),
        "pnl": _f(_cell(cons, f"H{tr}")),
        "return_pct": _f(_cell(cons, f"I{tr}")),
    }

    aum = {
        "total": _f(_cell(cons, f"C{_CONSOLIDATION_AUM_ROWS['total']}")),
        "schwab": {
            "value": _f(_cell(cons, f"B{_CONSOLIDATION_AUM_ROWS['schwab']}")),
            "pct": _f(_cell(cons, f"C{_CONSOLIDATION_AUM_ROWS['schwab']}")),
        },
        "etrade": {
            "value": _f(_cell(cons, f"B{_CONSOLIDATION_AUM_ROWS['etrade']}")),
            "pct": _f(_cell(cons, f"C{_CONSOLIDATION_AUM_ROWS['etrade']}")),
        },
    }

    # --- SW_Summary: returns-by-year (CCC-354 pool TWR, shared by all partners) ---
    by_year: list[dict[str, Any]] = []
    for r in _SW_YEAR_ROWS:
        period = _cell(swsum, f"A{r}")
        if not period:
            continue
        by_year.append({
            "period": str(period),
            "twr": _f(_cell(swsum, f"B{r}")),
            "partners": {
                name: _f(_cell(swsum, f"{col}{r}"))
                for col, name in _SW_YEAR_PARTNER_COLS.items()
            },
            "total_pnl": _f(_cell(swsum, f"H{r}")),
        })

    # --- ET_Summary: PM pool per-partner (Lien / CXZ) ---
    etrade_pool: list[dict[str, Any]] = []
    for r in _ET_SUMMARY_ROWS:
        name = _cell(etsum, f"A{r}")
        if not name:
            continue
        etrade_pool.append({
            "name": str(name),
            "contributed": _f(_cell(etsum, f"B{r}")),
            "cost_basis": _f(_cell(etsum, f"C{r}")),
            "current_value": _f(_cell(etsum, f"D{r}")),
            "share_pct": _f(_cell(etsum, f"E{r}")),
            "pool_pnl": _f(_cell(etsum, f"F{r}")),
            "return_on_invested": _f(_cell(etsum, f"G{r}")),
        })

    # --- SW_Summary: per-member CCC-354 detail (rows 5-9) keyed by partner name ---
    sw_members: dict[str, dict[str, Any]] = {}
    for r in _SW_MEMBER_ROWS:
        name = _cell(swsum, f"A{r}")
        if not name:
            continue
        sw_members[str(name)] = {
            "contributions": _f(_cell(swsum, f"B{r}")),
            "distributions": _f(_cell(swsum, f"C{r}")),
            "net_capital": _f(_cell(swsum, f"D{r}")),
            "pnl": _f(_cell(swsum, f"E{r}")),
            "balance": _f(_cell(swsum, f"F{r}")),
            "share_pct": _f(_cell(swsum, f"G{r}")),
            "simple_return": _f(_cell(swsum, f"H{r}")),
            "irr": _f(_cell(swsum, f"I{r}")),
        }

    # --- E*Trade per-member yearly returns (separate pre-merge accounts) ---
    def _et_years(ws, rows: dict[str, int]) -> list[dict[str, Any]]:
        out = []
        for period, r in rows.items():
            out.append({
                "period": period,
                "pnl": _f(_cell(ws, f"E{r}")),      # text (e.g. "数据不全") → None
                "return_pct": _f(_cell(ws, f"F{r}")),
            })
        return out

    etrade_by_year = {
        "lien": _et_years(lien_et, _ET_LIEN_YEAR_ROWS),
        "cxz": _et_years(cxz_et, _ET_CXZ_YEAR_ROWS),
    }

    # --- Member statements: join Consolidation + SW per-member + ET pool/years ---
    et_pool_by_name = {p["name"]: p for p in etrade_pool}
    member_statements = []
    for m in members:
        join = _MEMBER_JOIN.get(m["name"], {"sw": m["name"], "et": None})
        sw = sw_members.get(join["sw"]) if join["sw"] else None
        et_name = join["et"]
        ccc_by_year = [
            {"period": y["period"], "twr": y["twr"],
             "pnl": (y["partners"] or {}).get(_SW_YEAR_PARTNER_COLS_BY_SW.get(join["sw"], join["sw"]))}
            for y in by_year
        ] if join["sw"] else []
        et_block = None
        if et_name and et_name in et_pool_by_name:
            et_block = dict(et_pool_by_name[et_name])
            et_block["by_year"] = etrade_by_year.get(et_name.lower(), [])
        member_statements.append({
            "name": m["name"],
            "total": {
                "current": m["total_value"], "contrib": m["contrib"], "pnl": m["pnl"],
                "return_pct": m["return_pct"], "schwab_value": m["schwab_value"],
                "etrade_value": m["etrade_value"], "schwab_pct": m["schwab_pct"],
                "etrade_pct": m["etrade_pct"],
            },
            "ccc": ({**sw, "by_year": ccc_by_year} if sw else None),
            "etrade": et_block,
        })

    return {
        "members": members,
        "total": total,
        "aum": aum,
        "by_year": by_year,
        "etrade_pool": etrade_pool,
        "etrade_by_year": etrade_by_year,
        "member_statements": member_statements,
        "reconciled": {
            "schwab_ccc354": _latest_snapshot(swsnap),
            "etrade_pm": _latest_snapshot(etsnap),
        },
    }


def _latest_snapshot(ws) -> dict[str, Any] | None:
    """Return {date, value} for the last data row of a *_Snapshots sheet.

    Snapshot sheets are Date(A) | Total(B) | Note(C) | Closed?(D), data from row 5.
    Yellow placeholder rows have a note in C but no value in B — skip those.
    """
    latest = None
    max_row = ws.max_row or 0
    for r in range(5, max_row + 1):
        value = _f(_cell(ws, f"B{r}"))
        date_val = _cell(ws, f"A{r}")
        if value is None or date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.date().isoformat()
        else:
            date_str = str(date_val)[:10]
        latest = {"date": date_str, "value": value}
    return latest


def _mask_account(acct: str | None) -> str | None:
    """Show only the tail of an account id/hash so the PM can verify identity
    without exposing the full encrypted token."""
    if not acct:
        return None
    s = str(acct)
    return f"…{s[-4:]}" if len(s) > 4 else s


def live_nlv() -> dict[str, Any]:
    """Best-effort live broker NLV overlay vs the latest reconciled Excel snapshot.

    IMPORTANT (honesty guard): the live brokers expose whatever account the
    Schwab token / ETRADE_ACCOUNT_ID is bound to — which is NOT verified to be
    the partnership CCC-354 / PM pool. We therefore surface the masked account id
    and flag drift as informational only (``account_match: "unverified"``), never
    asserting the live account equals the Excel account.
    """
    book = read_book()
    reconciled = (book.get("reconciled") or {}) if book.get("available") else {}

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "schwab": _live_schwab(reconciled.get("schwab_ccc354")),
        "etrade": _live_etrade(reconciled.get("etrade_pm")),
    }


def _drift(live_val: float | None, recon: dict | None) -> dict[str, Any]:
    if live_val is None or not recon or recon.get("value") in (None, 0):
        return {"abs": None, "pct": None}
    base = recon["value"]
    return {"abs": live_val - base, "pct": (live_val - base) / base}


def _live_schwab(recon: dict | None) -> dict[str, Any]:
    out: dict[str, Any] = {
        "label": "Schwab",
        "reconciled": recon,           # {date, value} of latest CCC-354 snapshot
        "account_match": "unverified",  # live account not confirmed == CCC-354
    }
    try:
        from schwab import client as schwab_client  # type: ignore
    except Exception:
        out.update({"available": False, "message": "Schwab 模块不可用"})
        return out

    try:
        bal = schwab_client.get_account_balances()
    except Exception:
        log.warning("partnership_book: schwab balances failed", exc_info=True)
        out.update({"available": False, "message": "Schwab 余额读取失败"})
        return out

    if not bal.get("configured"):
        out.update({"available": False, "message": "Schwab 未配置"})
        return out
    if bal.get("stale") or not bal.get("authenticated"):
        out.update({"available": False, "message": "Schwab 需重新授权（数据陈旧）"})
        return out

    live_val = _f(bal.get("net_liquidation"))
    acct = None
    try:
        acct = schwab_client._account_number()  # type: ignore[attr-defined]
    except Exception:
        pass
    out.update({
        "available": True,
        "live_value": live_val,
        "account": _mask_account(acct),
        "updated_at": bal.get("updated_at"),
        "drift": _drift(live_val, recon),
    })
    return out


def _live_etrade(recon: dict | None) -> dict[str, Any]:
    out: dict[str, Any] = {
        "label": "E*Trade",
        "reconciled": recon,
        "account_match": "unverified",
    }
    try:
        from etrade import client as etrade_client  # type: ignore
        from etrade.auth import account_id as etrade_account_id  # type: ignore
    except Exception:
        out.update({"available": False, "message": "E*Trade 模块不可用"})
        return out

    try:
        bal = etrade_client.get_account_balances()
    except Exception:
        log.warning("partnership_book: etrade balances failed", exc_info=True)
        out.update({"available": False, "message": "E*Trade 余额读取失败"})
        return out

    if not bal.get("configured"):
        out.update({"available": False, "message": "E*Trade 未配置（ETRADE_ACCOUNT_ID 等）"})
        return out
    if bal.get("stale") or not bal.get("authenticated"):
        out.update({"available": False, "message": "E*Trade 需重新授权（数据陈旧）"})
        return out

    live_val = _f(bal.get("net_liquidation"))
    acct = None
    try:
        acct = etrade_account_id()
    except Exception:
        pass
    out.update({
        "available": True,
        "live_value": live_val,
        "account": _mask_account(acct),
        "updated_at": bal.get("updated_at"),
        "drift": _drift(live_val, recon),
    })
    return out
