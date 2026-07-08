"""SPEC-128 — one-shot migration: Partnership_Shares_v3.5.xlsx → data/book/.

Runs on the machine that has the workbook (the xlsx is git-ignored PII and
lives only on the dev Mac). Steps:
  1. extract every input sheet row → data/book/*.jsonl + config.json
  2. PARITY GATE: web.book_engine.compute_book() vs the workbook's own cached
     values (web.partnership_book oracle) — $ ±0.01, ratios ±0.0001,
     XIRR/TWR ±0.0005. Any mismatch → no .migrated marker, nonzero exit.
  3. anchor cross-check against v3.5 SPEC §9 key numbers
  4. write data/book/.migrated (oracle digest + timestamp)

Then: rsync data/book/ to oldair (deploy step, see SPEC-128 §5).
"""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import os
# Source workbook: prefer BOOK_XLSX env/argv (e.g. a FRESH export of the live
# Sheet) — 2026-07-06 lesson: the local copy was 5 weeks stale and the first
# migration missed a 6/26 contribution. Always migrate from a fresh export.
XLSX = Path(os.environ.get("BOOK_XLSX") or
            ROOT / "research" / "book_management" / "Partnership_Shares_v3.5.xlsx")
OUT = ROOT / "data" / "book"

CONFIG = {
    "sw_partners": ["Chinchaung", "Lien", "Xinzhong", "Alvin", "Lucas"],
    "et_partners": ["Lien", "CXZ"],
    # Consolidation display rows join SW + ET identities (Xinzhong ≡ CXZ)
    "members": [
        {"display": "Chinchaung",     "sw": "Chinchaung", "et": None},
        {"display": "Lien",           "sw": "Lien",       "et": "Lien"},
        {"display": "Xinzhong / CXZ", "sw": "Xinzhong",   "et": "CXZ"},
        {"display": "Alvin",          "sw": "Alvin",      "et": None},
        {"display": "Lucas",          "sw": "Lucas",      "et": None},
    ],
    "partner_display": {"Xinzhong": "Xinzhong/CXZ"},
    "subaccounts": ["CXZ-481", "Koching-276"],
    # §3.5 E*Trade merge (2026-06-01) dual-basis conventions
    "et_merge": {
        "date": "2026-06-01",
        # RESTATEMENT 2026-07-07 (PM ratified, restatements.jsonl): Lien basis
        # anchored 2024-11 (486,110.97 + net flows -269,553.99 = 216,556.98),
        # replacing option B 339,348.13 — pre-merge recorded gains +122,791.15
        # enter the book; ROI convention symmetric with CXZ.
        "cost_basis": {"Lien": 216556.98, "CXZ": 207907.0},
        "roi_convention": {
            "Lien": "value_vs_cost_basis",
            "CXZ": "value_vs_cost_basis",
        },
        "restated": "2026-07-07 anchor-2024-11 (see restatements.jsonl)",
    },
    "guarantees": [
        {"beneficiary": "Chinchaung", "guarantor": "Lien", "hurdle": 0.05,
         "period": "2025", "basis": 166181.0},
    ],
}

# v3.5 SPEC §9 anchors (cross-checked AFTER oracle parity — belt and braces)
ANCHORS = {
    "aum_total": 1265104.13,
    "sw_value": 627205.0,
    "et_value": 637899.13,
}


def _date_str(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    return str(v)[:10]


def _rows(ws, start: int, ncols: int):
    for r in range(start, (ws.max_row or 0) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ncols + 1)]
        if any(v is not None and str(v).strip() != "" for v in vals):
            yield r, vals


def extract() -> None:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)

    def dump(name: str, rows: list[dict]) -> None:
        p = OUT / name
        p.write_text("".join(json.dumps(r, ensure_ascii=False, sort_keys=True) + "\n"
                             for r in rows), encoding="utf-8")
        print(f"  {name}: {len(rows)} rows")

    # --- snapshots (Date | Total | Note | Closed?) from row 5 ---
    for sheet, fname, prefix in (("SW_Snapshots", "sw_snapshots.jsonl", "swsnap"),
                                 ("ET_Snapshots", "et_snapshots.jsonl", "etsnap")):
        ws = wb[sheet]
        rows = []
        for r, (d, total, note, closed) in _rows(ws, 5, 4):
            if d is None or total is None:
                continue   # yellow placeholder rows
            rows.append({"id": f"mig_{prefix}_{r:03d}", "event": "snapshot",
                         "date": _date_str(d), "total": float(total),
                         "note": str(note or ""), "closed": bool(closed)})
        dump(fname, rows)

    # --- cash ledgers (11 cols) from row 5 ---
    for sheet, fname, prefix in (("SW_CashLedger", "sw_cashledger.jsonl", "swflow"),
                                 ("ET_CashLedger", "et_cashledger.jsonl", "etflow")):
        ws = wb[sheet]
        rows = []
        for r, v in _rows(ws, 5, 11):
            snap_d, bank_d, partner, src, dst, amount, fee, instr, counts, ftype, note = v
            # Transit legs (bank → CXZ-481/Koching-276) legitimately carry NO
            # snapshot date — they don't belong to a capital period (Counts=No)
            # but they DO feed the sub-account pending math. Footer/total rows
            # have no Counts value → skip.
            if amount is None or counts is None:
                continue
            rows.append({"id": f"mig_{prefix}_{r:03d}", "event": "flow",
                         "snapshot_date": _date_str(snap_d) if snap_d else None,
                         "bank_date": _date_str(bank_d or snap_d) if (bank_d or snap_d) else None,
                         "partner": str(partner or ""), "from": str(src or ""),
                         "to": str(dst or ""), "amount": float(amount),
                         "fee": float(fee or 0.0), "instrument": str(instr or ""),
                         "counts": str(counts or "No").capitalize(),
                         "type": str(ftype or "").capitalize(),
                         "note": str(note or "")})
        dump(fname, rows)

    # --- personal series: marks + explicit year rows (verbatim conventions) ---
    cxz = wb["CXZ_ETrade"]
    rows = []
    for r, (month, value, *_rest) in _rows(cxz, 5, 2):
        if r >= 25 or value is None:   # year block starts at 25
            continue
        rows.append({"id": f"mig_cxzmark_{r:03d}", "event": "mark",
                     "month": str(month), "value": float(value)})
    for period, r in (("2024", 26), ("2025", 27), ("2026 YTD", 28)):
        start, end, net, _pnl = (cxz.cell(r, c).value for c in range(2, 6))
        rows.append({"id": f"mig_cxzyear_{r:03d}", "event": "year_row",
                     "period": period, "start": float(start), "end": float(end),
                     "net_flow": float(net or 0.0)})
    (lambda: None)()
    p = OUT / "cxz_etrade.jsonl"
    p.write_text("".join(json.dumps(x, ensure_ascii=False, sort_keys=True) + "\n"
                         for x in rows), encoding="utf-8")
    print(f"  cxz_etrade.jsonl: {len(rows)} rows")

    lien = wb["Lien_ETrade"]
    rows = []
    for r, (q, value, *_rest) in _rows(lien, 5, 2):
        if r >= 19 or value is None:
            continue
        rows.append({"id": f"mig_lienmark_{r:03d}", "event": "mark",
                     "month": str(q), "value": float(value)})
    # 2024 incomplete (statements only reach back to 2024-08) — explicit flag
    rows.append({"id": "mig_lienyear_020", "event": "year_row", "period": "2024",
                 "incomplete": True})
    for period, r in (("2025", 21), ("2026 YTD", 22)):
        start, end, net, _pnl = (lien.cell(r, c).value for c in range(2, 6))
        rows.append({"id": f"mig_lienyear_{r:03d}", "event": "year_row",
                     "period": period, "start": float(start), "end": float(end),
                     "net_flow": float(net or 0.0)})
    p = OUT / "lien_etrade.jsonl"
    p.write_text("".join(json.dumps(x, ensure_ascii=False, sort_keys=True) + "\n"
                         for x in rows), encoding="utf-8")
    print(f"  lien_etrade.jsonl: {len(rows)} rows")

    (OUT / "config.json").write_text(
        json.dumps(CONFIG, ensure_ascii=False, indent=1, sort_keys=True),
        encoding="utf-8")
    print("  config.json written")


# ── parity gate ───────────────────────────────────────────────────────────────

TOL_D = 0.01        # dollars
TOL_R = 1e-4        # ratios/shares (±0.01pp)
TOL_RET = 5e-4      # returns TWR/XIRR/simple (±0.05pp)


def _cmp(path: str, a, b, errs: list[str], tol_override: float | None = None):
    if a is None and b is None:
        return
    if (a is None) != (b is None):
        # oracle text cells (e.g. 数据不全) parse to None — allow None==None only
        errs.append(f"{path}: native={a!r} oracle={b!r}")
        return
    if isinstance(a, str) or isinstance(b, str):
        if str(a) != str(b):
            errs.append(f"{path}: native={a!r} oracle={b!r}")
        return
    tol = tol_override if tol_override is not None else (
        TOL_R if ("pct" in path or "share" in path) else
        TOL_RET if any(k in path for k in ("twr", "irr", "return")) else TOL_D)
    if abs(float(a) - float(b)) > tol:
        errs.append(f"{path}: native={a} oracle={b} (Δ={float(a) - float(b):+.6f})")


# Documented restatements: the archived workbook predates them; transform its
# values to post-restatement equivalents so parity stays exact. Return-ratio
# fields recompute from transformed components; convention-changed fields are
# excluded with a printed note.
RESTATEMENT_DELTA_LIEN = 339348.13 - 216556.98   # +122,791.15 pnl / -contrib


def _apply_restatements(oracle: dict) -> list[str]:
    notes = []
    d = RESTATEMENT_DELTA_LIEN
    for m in oracle.get("members", []):
        if m.get("name") == "Lien":
            m["contrib"] = (m["contrib"] or 0) - d
            m["pnl"] = (m["pnl"] or 0) + d
            m["return_pct"] = m["pnl"] / m["contrib"] if m["contrib"] else None
            notes.append("members[Lien] contrib/pnl/return transformed (restatement 2026-07-07)")
    tot = oracle.get("total", {})
    if tot:
        tot["contrib"] = (tot["contrib"] or 0) - d
        tot["pnl"] = (tot["pnl"] or 0) + d
        tot["return_pct"] = tot["pnl"] / tot["contrib"] if tot["contrib"] else None
    for row in oracle.get("etrade_pool", []):
        if row.get("name") == "Lien":
            row["cost_basis"] = 216556.98
            # ROI convention changed (TWR-since-2025 -> value_vs_cost_basis):
            # not delta-transformable, recompute directly
            cv = row.get("current_value") or 0
            row["return_on_invested"] = ((cv - 216556.98) / 216556.98) if cv else None
            notes.append("etrade_pool[Lien] basis/ROI transformed")
    # member_statements[].total mirrors members — same transform
    for s in oracle.get("member_statements", []):
        if s.get("name") == "Lien" and s.get("total"):
            st = s["total"]
            st["contrib"] = (st["contrib"] or 0) - d
            st["pnl"] = (st["pnl"] or 0) + d
            st["return_pct"] = st["pnl"] / st["contrib"] if st["contrib"] else None
    return notes


def parity(native: dict, oracle: dict) -> list[str]:
    errs: list[str] = []
    for n in _apply_restatements(oracle):
        print("  ~", n)
    # compare ORACLE keys only — the native payload may carry extra fields
    # (contributions_gross/distributions/dual_basis etc.) with no Excel
    # counterpart; the workbook is the reference, not a ceiling
    for i, (nm, om) in enumerate(zip(native["members"], oracle["members"])):
        for k in om:
            _cmp(f"members[{i}].{k}", nm.get(k), om.get(k), errs)
    for k in oracle["total"]:
        _cmp(f"total.{k}", native["total"].get(k), oracle["total"].get(k), errs)
    _cmp("aum.total", native["aum"]["total"], oracle["aum"]["total"], errs)
    for side in ("schwab", "etrade"):
        _cmp(f"aum.{side}.value", native["aum"][side]["value"], oracle["aum"][side]["value"], errs)
        _cmp(f"aum.{side}.pct", native["aum"][side]["pct"], oracle["aum"][side]["pct"], errs)
    for i, (ny, oy) in enumerate(zip(native["by_year"], oracle["by_year"])):
        _cmp(f"by_year[{i}].period", ny["period"], oy["period"], errs)
        _cmp(f"by_year[{i}].twr", ny["twr"], oy["twr"], errs)
        _cmp(f"by_year[{i}].total_pnl", ny["total_pnl"], oy["total_pnl"], errs)
        for pname, pv in (oy.get("partners") or {}).items():
            _cmp(f"by_year[{i}].partners.{pname}", (ny.get("partners") or {}).get(pname), pv, errs)
    for i, (np_, op_) in enumerate(zip(native["etrade_pool"], oracle["etrade_pool"])):
        for k in op_:
            _cmp(f"etrade_pool[{i}].{k}", np_.get(k), op_.get(k), errs)
    for who in ("lien", "cxz"):
        for i, oy in enumerate(oracle["etrade_by_year"][who]):
            ny = native["etrade_by_year"][who][i]
            _cmp(f"etrade_by_year.{who}[{i}].pnl", ny.get("pnl"), oy.get("pnl"), errs)
            _cmp(f"etrade_by_year.{who}[{i}].return_pct", ny.get("return_pct"), oy.get("return_pct"), errs)
    for i, (ns, os_) in enumerate(zip(native["member_statements"], oracle["member_statements"])):
        for k, v in (os_.get("ccc") or {}).items():
            if k == "by_year":
                continue
            _cmp(f"stmt[{i}].ccc.{k}", (ns.get("ccc") or {}).get(k), v, errs)
    if len(native["members"]) != len(oracle["members"]):
        errs.append("member count mismatch")
    return errs


def main() -> int:
    assert XLSX.exists(), f"workbook missing: {XLSX}"
    print(f"source workbook: {XLSX}")

    # preserve rows recorded through the native UI (they carry recorded_at) —
    # re-extraction rewrites the migrated base, then these are re-appended
    preserved: dict[str, list[str]] = {}
    for f in OUT.glob("*.jsonl"):
        keep = [l for l in f.read_text(encoding="utf-8").splitlines()
                if l.strip() and '"recorded_at"' in l]
        if keep:
            preserved[f.name] = keep
            print(f"  preserving {len(keep)} native rows in {f.name}")

    print("extracting inputs from workbook…")
    extract()

    def _reappend():
        for fname, lines in preserved.items():
            with (OUT / fname).open("a", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")
            print(f"  re-appended {len(lines)} native rows to {fname}")
        preserved.clear()

    print("running parity gate (native engine vs workbook cached values)…")
    from web.book_engine import compute_book
    from web.partnership_book import _parse_bytes_or_path
    native = compute_book(OUT, force=True)
    assert native.get("available"), native
    oracle = _parse_bytes_or_path(XLSX)
    assert oracle is not None, "oracle parse failed"

    errs = parity(native, oracle)
    if errs:
        print(f"\nPARITY FAIL — {len(errs)} mismatches (no .migrated marker):")
        for e in errs[:40]:
            print("  ✗", e)
        _reappend()   # never strand natively-recorded rows on a failed gate
        return 1
    print("parity: ALL FIELDS MATCH (tolerances $0.01 / 0.01pp / 0.05pp)")

    # AUM equality with the oracle is already asserted field-by-field in
    # parity(); the v3.5 §9 constants were only valid for the first migration
    # (2026-07-06 lesson: a FRESH sheet has newer snapshots and the hardcoded
    # anchors go stale). Print for human eyeball instead of hard-gating.
    a = native["aum"]
    print(f"AUM (oracle-matched): total={a['total']:,.2f} "
          f"SW={a['schwab']['value']:,.2f} ET={a['etrade']['value']:,.2f}")

    if not native.get("recon_all_green"):
        print("recon checks not all green:")
        for c in native["recon_checks"]:
            if not c["ok"]:
                print("  ✗", c["name"], c["detail"])
        _reappend()
        return 1
    print(f"recon checks: {len(native['recon_checks'])}/{len(native['recon_checks'])} green")

    # re-append preserved native rows AFTER the parity gate (parity must be
    # judged against the workbook-equivalent state)
    _reappend()

    digest = hashlib.sha256(json.dumps(
        {k: oracle[k] for k in ("members", "total", "aum")},
        sort_keys=True, default=str).encode()).hexdigest()[:16]
    (OUT / ".migrated").write_text(json.dumps({
        "at": datetime.now().isoformat(timespec="seconds"),
        "oracle_digest": digest,
        "workbook": str(XLSX.name),
    }, indent=1))
    print(f"\nMIGRATED ✓  oracle digest {digest}")
    print("next: rsync data/book/ to oldair (SPEC-128 §5), then PM renames the "
          "Drive sheet to *_archive")
    return 0


if __name__ == "__main__":
    sys.exit(main())
